from openpyxl import load_workbook
import pytest

from exporters.excel_exporter import ExcelExporter


def test_excel_exporter_creates_multiple_sheets(tmp_path):
    output_file = tmp_path / "output.xlsx"

    data = {
        "Tokyo": [
            {"a": 1, "b": 2},
            {"a": 3, "b": 4},
        ],
        "London": [
            {"a": 5, "b": 6},
        ],
    }

    exporter = ExcelExporter(output_file)
    exporter.export(data)

    assert output_file.exists()

    wb = load_workbook(output_file)
    assert set(wb.sheetnames) == {"Tokyo", "London"}


def test_excel_exporter_rejects_invalid_input(tmp_path):
    exporter = ExcelExporter(tmp_path / "out.xlsx")

    with pytest.raises(TypeError):
        exporter.export(["not", "a", "dict"])
