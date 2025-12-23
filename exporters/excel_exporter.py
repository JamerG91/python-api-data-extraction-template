from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .base import Exporter


class ExcelExporter(Exporter):
    def __init__(self, output_path: str):
        self.output_path = output_path

    def export_single(self, data: list[dict], sheet_name: str = "data"):
        if not data:
            return

        if not isinstance(data, list) or not isinstance(data[0], dict):
            raise ValueError("ExcelExporter expects a list of dictionaries")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        self._write_sheet(ws, data)
        wb.save(self.output_path)

    def export(self, sheets: dict[str, list[dict]]):
        """
        Export multiple datasets into a single Excel workbook, using one sheet
        per dataset.

        Each key in `sheets` is used as the worksheet name (truncated to Excel's
        31-character limit), and each value must be a list of dictionaries
        representing tabular rows.

        Parameters
        ----------
        sheets : dict[str, list[dict]]
            Mapping of sheet names to row data. Each value must be a list of
            dictionaries with identical keys.

        Raises
        ------
        TypeError
            If `sheets` is not a dictionary, or if any value is not a list of
            dictionaries.
        ValueError
            If `sheets` is empty.
        """

        if not isinstance(sheets, dict):
            raise TypeError(
                "sheets must be a dictionary mapping sheet names to list of dicts"
            )

        if not sheets:
            raise ValueError("sheets dictionary is empty")

        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        for sheet_name, data in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit
            self._write_sheet(ws, data)

        wb.save(self.output_path)

    def _write_sheet(self, ws, data: list[dict]):
        if not data:
            return

        if data and not isinstance(data[0], dict):
            raise TypeError(
                "Expected data to be a list of dictionaries, "
                f"got list of {type(data[0]).__name__}"
            )

        headers = list(data[0].keys())
        ws.append(headers)

        for row in data:
            ws.append([row.get(h) for h in headers])

        self._autosize(ws, len(headers))

    def _autosize(self, ws, column_count: int):
        for i in range(1, column_count + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
