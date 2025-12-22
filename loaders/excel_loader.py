from openpyxl import load_workbook
from .base import Loader


class ExcelLoader(Loader):
    def __init__(self, path: str, sheet_name=None):
        self.path = path
        self.sheet_name = sheet_name

    def load(self):
        wb = load_workbook(self.path)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active

        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0]

        return [dict(zip(headers, row)) for row in rows[1:]]
